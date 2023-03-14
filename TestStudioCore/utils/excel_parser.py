# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import os
import sys


class ReadExcel(object):
    def __init__(self, file_path, sheet_name=""):

        self.sheet_name = sheet_name
        self.file_path = file_path
        if not os.path.exists(self.file_path):
            sys.exit(1)
        try:
            self.workbook = load_workbook(self.file_path)
        except Exception as e:
            print("打开文件发生错误：%s" % e)


        self.ws = self.workbook["测试用例"]
        # 表头
        self.title = list(self.ws.rows)[0]
        # 获取sheet的最大行数和列数
        self.rows = self.ws.max_row
        self.cols = self.ws.max_column
        if self.rows == 0 or self.cols == 0:
            print("表格为空，请确认测试数据是否正常写入！")
            # 有错误退出
            sys.exit(1)

    def readExcel_to_dict(self):
        s_data = []
        for i in range(2, self.rows + 1):
            sheet_data = {}
            for j in range(1, self.cols + 1):
                c_data = self.ws.cell(i, j).value
                # 这里-1 目的 因为excel是从1开始，python列表需要从0开始取值
                sheet_data[self.title[j - 1].value] = c_data
            if sheet_data["层级"] is not None:
                s_data.append(sheet_data)
        return s_data

